from copomap import copo
import openpyxl
import xlrd
from pathlib import Path
import xlwt
from openpyxl import load_workbook
loc = ("1.xlsx");
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
wb2 = load_workbook (loc)
sheets=wb2.sheetnames
Sheet1=wb2 [sheets[0]]

for j in range(sheet.nrows-1):
    for i in range(sheet.ncols-3):
    ##    print(sheet.nrows)
    ##    print(sheet.ncols)
        CO1=(sheet.cell_value(j+1, 2))
        PO1=(sheet.cell_value(0, i+3))
        ans=copo(CO1,PO1);
        if ans >0.29:
            ans=3;
        if ans >0.19:
            ans=2;
        if ans >0.09:
            ans=1
        if ans <= 0.09:
            ans=" "
            
        print(j)
        print(i)
        print(PO1)
        print(CO1)
        print(ans)
        Sheet1.cell(row=j+2,column=i+4).value =ans
    wb2.save ("1.xlsx")
